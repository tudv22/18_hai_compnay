import base64

from odoo import fields, api, models
from odoo.exceptions import UserError

from io import BytesIO
from openpyxl import load_workbook
from openpyxl.styles import Font, borders, Alignment
from datetime import date, datetime, time


class InOutStockReport(models.TransientModel):
    _name = 'in.out.stock.report'
    _description = 'In/Out Stock Report'

    today = date.today()
    month = today.month if today.month != 0 else 1
    year = today.year
    first_day_of_month = date(year, month, 1)

    company_id = fields.Many2one(
        'res.company',
        string='Công ty',
        default=lambda self: self.env.user.company_id.id,
        domain=lambda self: [('id', 'in', self.env.user.company_ids.ids)]
    )
    warehouse_ids = fields.Many2many(
        'stock.warehouse',
        string='Kho cần truy xuất',
        domain="[('company_id', '=', company_id)]",
    )
    start_date = fields.Date(
        string='Start date',
        default=first_day_of_month
    )
    end_date = fields.Date(
        string='End date',
        default=today
    )

    @api.onchange('company_id')
    def _onchange_warehouse_id(self):
        for record in self:
            record.warehouse_ids = False

    def get_data(self, warehouse_id):
        query = """
                WITH location_data AS ( -- Lấy các location của kho, company đã chọn
                        SELECT id, name
                        FROM stock_location
                        WHERE company_id = {company_id}
                        AND usage = 'internal'
                        AND warehouse_id = {warehouse_id}
                    ),
                    stock_quant_data AS ( -- Số lượng tồn hiện tại
                        SELECT sq.product_id, SUM(sq.quantity) AS total_quantity
                        FROM stock_quant sq
                        WHERE sq.company_id = {company_id}
                        AND (sq.location_id IN (SELECT id FROM location_data))
                        GROUP BY sq.product_id
                    ),
                    stock_move_quantity_end_current AS ( -- Số lượng nhập xuất từ cuối kì đến hiện tại
                        SELECT 
                            smq.product_id,
                            SUM(CASE WHEN smq.location_dest_id IN (SELECT id FROM location_data) THEN smq.quantity ELSE 0 END) AS quantity_in,
                            SUM(CASE WHEN smq.location_id IN (SELECT id FROM location_data) THEN smq.quantity ELSE 0 END) AS quantity_out
                        FROM stock_move smq
                        LEFT JOIN stock_picking sp ON smq.picking_id = sp.id
                        WHERE smq.company_id = {company_id}
                        AND sp.date_done > '{end_date}' AND sp.date_done <= '{today}'
                        AND (smq.location_id IN (SELECT id FROM location_data) OR smq.location_dest_id IN (SELECT id FROM location_data))
                        GROUP BY smq.product_id
                    ),
                    stock_move_quantity_data AS ( -- Số lượng nhập và xuất trong kì start - end
                        SELECT 
                            smq.product_id,
                            SUM(CASE WHEN smq.location_dest_id IN (SELECT id FROM location_data) THEN smq.quantity ELSE 0 END) AS quantity_in,
                            SUM(CASE WHEN smq.location_id IN (SELECT id FROM location_data) THEN smq.quantity ELSE 0 END) AS quantity_out
                        FROM stock_move smq
                        LEFT JOIN stock_picking sp ON smq.picking_id = sp.id
                        WHERE smq.company_id = {company_id}
                        AND sp.date_done BETWEEN '{start_date}' AND '{end_date}'
                        AND (smq.location_id IN (SELECT id FROM location_data) OR smq.location_dest_id IN (SELECT id FROM location_data))
                        GROUP BY smq.product_id
                    ),
                    sh_warehouse_cost_data AS ( -- Giá mỗi sản phẩm
                        SELECT swc.product_id, MAX(swc.cost) AS cost
                        FROM sh_warehouse_cost swc
                        GROUP BY swc.product_id
                    ),
                    stock_move_value_end_current AS (-- Giá trị nhập xuất từ cuối kì đến hiện tại
                        SELECT 
                            sm.product_id,
                            ABS(SUM(CASE WHEN sm.location_dest_id IN (SELECT id FROM location_data) THEN svl.value ELSE 0 END)) AS value_in,
                            ABS(SUM(CASE WHEN sm.location_id IN (SELECT id FROM location_data) THEN svl.value ELSE 0 END)) AS value_out
                        FROM stock_move sm
                        LEFT JOIN stock_valuation_layer svl ON sm.id = svl.stock_move_id
                        LEFT JOIN stock_picking sp ON sm.picking_id = sp.id
                        WHERE sm.company_id = {company_id}
                        AND sp.date_done > '{end_date}' AND sp.date_done <= '{today}'
                        AND (sm.location_id IN (SELECT id FROM location_data) OR sm.location_dest_id IN (SELECT id FROM location_data))
                        GROUP BY sm.product_id
                    ),
                    stock_move_value_data AS (-- Giá trị nhập xuất trong kì start - end
                         SELECT 
                            sm.product_id,
                            ABS(SUM(CASE WHEN sm.location_dest_id IN (SELECT id FROM location_data) THEN svl.value ELSE 0 END)) AS value_in,
                            ABS(SUM(CASE WHEN sm.location_id IN (SELECT id FROM location_data) THEN svl.value ELSE 0 END)) AS value_out
                        FROM stock_move sm
                        LEFT JOIN stock_valuation_layer svl ON sm.id = svl.stock_move_id
                        LEFT JOIN stock_picking sp ON sm.picking_id = sp.id
                        WHERE sm.company_id = {company_id}
                        AND sp.date_done BETWEEN '{start_date}' AND '{end_date}'
                        AND (sm.location_id IN (SELECT id FROM location_data) OR sm.location_dest_id IN (SELECT id FROM location_data))
                        GROUP BY sm.product_id
                    )
                    SELECT 
                        pp.id AS product_id,
                        pp.default_code AS default_code,
                        (COALESCE(sq.total_quantity, 0) - COALESCE(smqec.quantity_in, 0) + COALESCE(smqec.quantity_out, 0)) AS end_quantity,
                        (COALESCE(sq.total_quantity, 0) - COALESCE(smqec.quantity_in, 0) + COALESCE(smqec.quantity_out, 0)) * COALESCE(swc.cost, 0) AS end_value,
                        COALESCE(smq.quantity_in, 0) AS quantity_in,
                        COALESCE(smq.quantity_out, 0) AS quantity_out,
                        COALESCE(smv.value_in, 0) AS value_in,
                        COALESCE(smv.value_out, 0) AS value_out,
                        (COALESCE(sq.total_quantity, 0) - COALESCE(smqec.quantity_in, 0) + COALESCE(smqec.quantity_out, 0) - COALESCE(smq.quantity_in, 0) + COALESCE(smq.quantity_out, 0)) AS start_quantity,
                        (COALESCE(sq.total_quantity, 0) - COALESCE(smqec.quantity_in, 0) + COALESCE(smqec.quantity_out, 0)) * COALESCE(swc.cost, 0)
                            - COALESCE(smv.value_in, 0)
                            + COALESCE(smv.value_out, 0) AS start_value
                    FROM stock_quant_data sq
                    LEFT JOIN product_product pp ON sq.product_id = pp.id
                    LEFT JOIN stock_move_quantity_end_current smqec on sq.product_id = smqec.product_id
                    LEFT JOIN stock_move_quantity_data smq on sq.product_id = smq.product_id
                    LEFT JOIN sh_warehouse_cost_data swc on sq.product_id = swc.product_id
                    LEFT JOIN stock_move_value_end_current smvec on sq.product_id = smvec.product_id
                    LEFT JOIN stock_move_value_data smv on sq.product_id = smv.product_id
                    
        """.format(company_id=self.company_id.id, warehouse_id=warehouse_id.id,
                   start_date=str(self.start_date), today=self.today,
                   end_date=datetime.combine(self.end_date, time(23, 59, 59)).strftime('%Y-%m-%d %H:%M:%S'))
        self._cr.execute(query)
        datas = self._cr.dictfetchall()
        excel_data = []

        total_end_quantity = 0
        total_end_value = 0
        total_quantity_in = 0
        total_quantity_out = 0
        total_value_in = 0
        total_value_out = 0
        total_start_quantity = 0
        total_start_value = 0

        product_ids = [row["product_id"] for row in datas]
        products = self.env["product.product"].browse(product_ids)
        product_mapping = {}
        for product in products:
            # Lấy danh sách thuộc tính của sản phẩm
            attribute_values = product.product_template_attribute_value_ids.mapped("name")
            attribute_suffix = f" ({', '.join(attribute_values)})" if attribute_values else ""

            # Gán tên sản phẩm kèm theo thuộc tính
            product_mapping[product.id] = {
                "name": f"{product.name}{attribute_suffix}",
                "uom_id": product.uom_id.name,
                "categ_id": product.categ_id.complete_name
            }
        for row in datas:
            # Thêm tên sản phẩm vào dòng dữ liệu
            product_id = row.get("product_id")
            product_data = product_mapping.get(product_id, {})
            row["product_name"] = product_data.get("name")
            row["uom_id"] = product_data.get("uom_id")
            row["category_id"] = product_data.get("categ_id")
            # Tính tổng các cột
            total_end_quantity += row.get("end_quantity", 0)
            total_end_value += row.get("end_value", 0)
            total_quantity_in += row.get("quantity_in", 0)
            total_quantity_out += row.get("quantity_out", 0)
            total_value_in += row.get("value_in", 0)
            total_value_out += row.get("value_out", 0)
            total_start_quantity += row.get("start_quantity", 0)
            total_start_value += row.get("start_value", 0)
            excel_data.append(row)

        # Thêm hàng tổng vào đầu danh sách
        total_row = {
            "default_code": warehouse_id.name,
            "end_quantity": total_end_quantity,
            "end_value": total_end_value,
            "quantity_in": total_quantity_in,
            "quantity_out": total_quantity_out,
            "value_in": total_value_in,
            "value_out": total_value_out,
            "start_quantity": total_start_quantity,
            "start_value": total_start_value,
        }
        excel_data.insert(0, total_row)

        return excel_data

    def create_warehouse_report(self, row, warehouse_id, ws, column_mapping, datas):
        for index, line in enumerate(datas):
            if index == 0:  # Dòng tổng
                merge_range = f"A{row}:D{row}"  # Merge 4 ô đầu
                ws.merge_cells(merge_range)
                ws[f"A{row}"].value = line.get("default_code", "")  # Ghi giá trị tổng
                ws[f"A{row}"].font = Font(name='Times New Roman', size=12, bold=True)
                ws[f"A{row}"].alignment = Alignment(horizontal="center", vertical="center")

                # Ghi các ô còn lại
                for key, col in column_mapping.items():
                    if key not in ["default_code", "product_name", "uom_id", "category_id"]:  # Bỏ qua 4 cột đầu đã merge
                        ws[f"{col}{row}"].font = Font(name='Times New Roman', size=12, bold=True)
                        ws[f"{col}{row}"].value = line.get(key)
                        if isinstance(line.get(key), (int, float)):
                            ws[f"{col}{row}"].number_format = "#,##0"
            else:
                for key, col in column_mapping.items():
                    ws[f"{col}{row}"].font = Font(name='Times New Roman', size=12)
                    ws[f"{col}{row}"].value = line.get(key)
                    if isinstance(line.get(key), (int, float)):
                        ws[f"{col}{row}"].number_format = "#,##0"
            row += 1
        row += 2
        return row

    def create_report_quantity(self):
        url = f'/stock_quantity/report/excel?report_id={self.id}'
        return {
            'type': 'ir.actions.act_url',
            'url': url,
            'target': 'new'
        }

    def get_warehouse(self):
        self.env.registry.clear_cache()
        warehouse_ids = self.warehouse_ids
        if not warehouse_ids:
            return self.warehouse_ids.search([('company_id', '=', self.company_id.id)])
        return warehouse_ids