# -*- coding: utf-8 -*-
import base64
import urllib.parse

from odoo import http
from odoo.exceptions import UserError
from odoo.http import request
from io import BytesIO
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment


class IOStockReportDownloadExcel(http.Controller):

    @http.route('/stock_quantity/report/excel', type='http', auth="public")
    def download_stock_quantity_report_excel(self, **kwargs):
        report = request.env['in.out.stock.report'].browse(int(kwargs.get('report_id')))
        template = report.env['ir.attachment'].browse(report.env.ref('a1_stock_report.IO_stock_report_quantity_attachment').id)
        if not template or not template.datas:
            raise UserError("Không tìm thấy tệp báo cáo mẫu!")

        decode = base64.b64decode(template.datas)
        wb = load_workbook(BytesIO(decode))
        ws = wb.active

        header_font = Font(name='Times New Roman', size=12, bold=True)
        header_aligment = Alignment(horizontal='left', vertical='center')
        ws['A1'].value = 'Đơn vị: ' + report.company_id.name
        ws['A2'].value = 'Địa chỉ: ' + report.company_id.partner_id.street
        for cell in ['A1', 'A2']:
            ws[cell].font = header_font
            ws[cell].alignment = header_aligment
        date_range = 'Từ ngày: %s đến ngày: %s' % (
            report.start_date.strftime('%d/%m/%Y'), report.end_date.strftime('%d/%m/%Y')
        )
        ws.merge_cells('A5:H5')
        ws['A5'].value = date_range
        ws['A5'].font = Font(name='Times New Roman', size=12)
        ws['A5'].alignment = Alignment(horizontal='center', vertical='center')
        column_mapping = {
            "default_code": "A",
            "product_name": "B",
            "category_id": "C",
            "uom_id": "D",
            "start_quantity": "E",
            "quantity_in": "F",
            "quantity_out": "G",
            "end_quantity": "H",
        }

        row = 10
        for warehouse_id in report.get_warehouse():
            datas = report.get_data(warehouse_id)
            row = report.create_warehouse_report(row, warehouse_id, ws, column_mapping, datas)

        column_widths = {
            "A": 30,  # default_code
            "B": 35,  # product_name
            "C": 35,  # category_id
            "D": 10,  # uom_id
            "E": 12,  # start_quantity
            "F": 12,  # quantity_in
            "G": 12,  # quantity_out
            "H": 12,  # end_quantity
        }
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width

        fp = BytesIO()
        wb.save(fp)
        fp.seek(0)

        file_content = fp.read()
        fp.close()
        file_name = 'Báo cáo XNT kho (Số lượng).xlsx'
        file_name = urllib.parse.quote(file_name)

        respone = self.respone_download(file_content, file_name)
        return respone

    def respone_download(self, file_content, file_name):
        respone = request.make_response(
            file_content,
            headers=[
                ('Content-Type', 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'),
                ('Content-Disposition', f'attachment; filename="{file_name}"'),
            ]
        )
        return respone